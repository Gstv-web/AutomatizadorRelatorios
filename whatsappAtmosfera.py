import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import rows_from_range
import os


pathEnviados = "./Entrada/Atmosfera/Enviados/"
pathRecebidos = "./Entrada/Atmosfera/Recebidos/"

def whatsappAtmosfera():

    arquivosEnviados = os.listdir(pathEnviados)
    arquivosRecebidos = os.listdir(pathRecebidos)

    arquivosEnviados.sort()
    arquivosRecebidos.sort()

    for i, file in enumerate(arquivosEnviados):
    # LER AQUI O ENVIADO
        readerEnviados = pd.read_excel(pathEnviados + file, sheet_name="RELATÓRIO", skiprows=0)
    # ESSE AQUI É O NOME DO ARQUIVO RECEBIDO PRA ABIR NA LINHA DE BAIXO
        
        readerRecebidos = pd.read_excel(pathRecebidos + arquivosRecebidos[i], sheet_name="RESPOSTAS", skiprows=0)
        
    # ler arquivos

    
    # readerEnviados = pd.read_excel('Enviados Fornecedor B.xlsx', sheet_name='RELATÓRIO', skiprows=0)
    # readerRecebidos = pd.read_excel('Recebidos Fornecedor B.xlsx', sheet_name='RESPOSTAS', skiprows=0)

    # Fazendo uma copia das planilhas para não alterar o original
        enviados_df = readerEnviados.copy()
        recebidos_df = readerRecebidos.copy()



        # EDITANDO PLANILHA "ENVIADOS"
        # Removendo coluna 1, linhas vazias e renomeando colunas necessárias, remoção de '55' dos telefones e remoção de duplicatas
        # enviados_df = pd.read_excel('Enviados Fornecedor B.xlsx', sheet_name='RELATÓRIO')
        enviados_df.drop('Unnamed: 0', inplace=True, axis=1)
        enviados_df.drop(enviados_df.index[[0,1,2,3,4,5]], inplace=True, axis=0)
        enviados_df.rename(columns={'Unnamed: 1': 'TELEFONE', 'Unnamed: 2' : 'DATA', 'Unnamed: 7' : 'STATUS'}, inplace=True)

        for index in enviados_df.index: 
            numero = str(enviados_df.loc[index, 'TELEFONE'])

            if numero[0:2] == '55':
                novoNum = numero.lstrip('55')
            
            enviados_df.at[index, 'TELEFONE'] = novoNum

        enviados_df.drop_duplicates(subset=['TELEFONE'], keep='first', inplace=True)

        enviados_df.drop(enviados_df.columns[[2,3,4,5]], inplace=True, axis=1)

        enviados_df = enviados_df[['TELEFONE', 'DATA', 'STATUS']]


        # EDITANDO PLANILHA "RESPOSTAS"
        # recebidos_df = pd.read_excel('Recebidos Fornecedor B.xlsx', sheet_name='RESPOSTAS')
        recebidos_df.drop('Unnamed: 0', inplace=True, axis=1)
        recebidos_df.drop(recebidos_df.index[[0,1,2,3,4,5]], inplace=True, axis=0)
        recebidos_df.rename(columns={'Unnamed: 1' : 'TELEFONE', 'Unnamed: 2' : 'MENSAGEM'}, inplace=True)


        for index in recebidos_df.index: 
            numero = str(recebidos_df.loc[index, 'TELEFONE'])

            if numero[0:2] == '55':
                novoNum = numero.lstrip('55')
            
            recebidos_df.at[index, 'TELEFONE'] = novoNum

        # Apesar de haver duplicadas, há mensagens diferentes e talvez possa ser usada, caso contrário, remover
        # recebidos_df.drop_duplicates(subset=['TELEFONE'], keep='first', inplace=True)

        recebidos_df.drop(recebidos_df.columns[[2,3,4,5,6,7]], inplace=True, axis=1)


        # remover linhas “no” alterar e células de “yes” para “enviado” 
        for index in enviados_df.index:
            if enviados_df.loc[index, 'STATUS'] == 'No':
                enviados_df.drop(index, inplace=True, axis=0)

        for index in enviados_df.index:
            newValue = str('Enviado')
            if enviados_df.loc[index, 'STATUS'] == 'Yes':
                enviados_df.at[index, 'STATUS'] = newValue



        # OPENPYXL
        with pd.ExcelWriter('./Conversoes/FornecedorB.xlsx') as writer:
            enviados_df.to_excel(writer, sheet_name='Enviados', index=False)
            recebidos_df.to_excel(writer, sheet_name='Respostas', index=False)



        reader = load_workbook('./Conversoes/FornecedorB.xlsx')


        enviados_ws = reader['Enviados']
        recebidos_ws = reader['Respostas']


        enviados_ws.column_dimensions['A'].width = 17
        enviados_ws.column_dimensions['B'].width = 14
        enviados_ws.column_dimensions['C'].width = 14

        recebidos_ws.column_dimensions['A'].width = 17
        recebidos_ws.column_dimensions['B'].width = 200

        for rows in enviados_ws.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')

        for rows in recebidos_ws.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')


        for col in enviados_ws.iter_cols(min_row=2, min_col=2, max_col=2):
            for cell in col:
                celula = cell.value.strftime("%d/%m/%Y") # aqui eu acessei o valor da célula, agora eu posso alterar o valor
                dateSet = celula.split() # O estilo de data é diferente e precisa ser alterado
                justDate = dateSet[0] # Aqui eu acesso apenas a data
                cell.value = justDate # Aqui eu altero o valor da célula



        # Bordas das células

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        for col in enviados_ws.iter_cols(min_row=2, min_col=1, max_col=3):
            for cell in col:
                cell.border = thin_border


        for col in recebidos_ws.iter_cols(min_row=2, min_col=1, max_col=2):
            for cell in col:
                cell.border = thin_border


        reader.save("./Saida/Atmosfera/RelatorioTeste.xlsx")

