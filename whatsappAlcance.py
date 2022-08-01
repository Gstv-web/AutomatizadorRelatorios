import pandas as pd
import numpy as np
import xlsxwriter as xlsx
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import os


path = './Entrada/Alcance/'
pathSaida = './Saida/Alcance/'
blacklist = pd.read_csv('./Blacklist.csv')


def whatsappAlcance():

    counter = 0


    arquivosEnviados = os.listdir(path)

    arquivosEnviados.sort()

    for i, file in enumerate(arquivosEnviados):
        reader = pd.read_csv(path + file, sep=';')

        reader.rename(columns={'contato' : 'TELEFONE', 'dt_envio' : 'DATA', 'resp' : 'MENSAGEM'}, inplace=True)



        enviados_df = reader.copy()
        recebidos_df = reader.copy()


        #Inserindo coluna "Status"
        enviados_df.insert(3, 'STATUS', 'Enviado')



        # RELATÓRIO


        for index in enviados_df.index:
            numero = str(enviados_df.loc[index, 'TELEFONE'])

            if numero[0:2] == '55':
                novoNum = numero.lstrip('55')

            enviados_df.at[index, 'TELEFONE'] = novoNum


        enviados_df.drop_duplicates(subset=['TELEFONE'], keep='first', inplace=True)

        enviados_df.drop(enviados_df.columns[2], inplace=True, axis=1)

        #Removendo linhas com erro de envio
        for index in enviados_df.index:
            if enviados_df.loc[index, 'DATA'] == "error":
                enviados_df.drop(index, inplace=True, axis=0)

        # Removendo números da blacklist
        numBlacklist = []
        for index in blacklist.index:
            num = str(blacklist.loc[index, 'Telefone'])

            novoNum = num
            blacklist.at[index, 'Telefone'] = novoNum

            numBlacklist.append(blacklist.loc[index, 'Telefone'])

        for index in enviados_df.index:
            if enviados_df.loc[index, 'TELEFONE'] in numBlacklist:
                enviados_df.drop(index, inplace=True, axis=0)



        # RESPOSTAS


        recebidos_df.drop(recebidos_df.columns[1], inplace=True, axis=1)


        for index in recebidos_df.index:
            numero = str(recebidos_df.loc[index, 'TELEFONE'])

            if numero[0:2] == '55':
                novoNum = numero.lstrip('55')

            recebidos_df.at[index, 'TELEFONE'] = novoNum


        for index in recebidos_df.index:
            status = str(recebidos_df.loc[index, 'MENSAGEM'])
            if recebidos_df.loc[index, 'MENSAGEM'] != "SAIR DA LISTA" and recebidos_df.loc[index, 'MENSAGEM'] != "SIM":
                recebidos_df.drop(index, inplace=True, axis=0)



        # OPENPYXL

        counter += 1
        strCounter = str(counter)
        with pd.ExcelWriter(f'./{strCounter}-file.xlsx') as writer:
            enviados_df.to_excel(writer, sheet_name='Enviados', index=False)
            recebidos_df.to_excel(writer, sheet_name='Respostas', index=False)


        excelReader = load_workbook(f'./{strCounter}-file.xlsx')

        enviados_ws = excelReader['Enviados']
        recebidos_ws = excelReader['Respostas']

        enviados_ws.column_dimensions['A'].width = 17
        enviados_ws.column_dimensions['B'].width = 14
        enviados_ws.column_dimensions['C'].width = 14

        recebidos_ws.column_dimensions['A'].width = 17
        recebidos_ws.column_dimensions['B'].width = 200

        for rows in enviados_ws.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')

        for rows in recebidos_ws.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')


        for col in enviados_ws.iter_cols(min_row=2, min_col=2, max_col=2):
            for cell in col:
                celula = cell.value# aqui eu acessei o valor da célula, agora eu posso alterar o valor
                dateSet = celula.split() # O estilo de data é diferente e precisa ser alterado
                justDate = dateSet[0] # Aqui eu acesso apenas a data
                cell.value = justDate # Aqui eu altero o valor da célula


        # Bordas das células

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        for col in enviados_ws.iter_cols(min_row=2, min_col=1, max_col=3):
            for cell in col:
                cell.border = thin_border


        for col in recebidos_ws.iter_cols(min_row=2, min_col=1, max_col=2):
            for cell in col:
                cell.border = thin_border


        excelReader.save(f"{pathSaida}{strCounter}-result.xlsx")