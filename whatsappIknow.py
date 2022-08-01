import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import os


pathEnviados = "./Entrada/iKnow/Enviados/"
pathRecebidos = "./Entrada/iKnow/Recebidos/"
blacklist = pd.read_csv('./Blacklist.csv')
pathSaida = "./Saida/iKnow/"

def whatsappIknow():
    
    counter = 0

    arquivosEnviados = os.listdir(pathEnviados)
    arquivosRecebidos = os.listdir(pathRecebidos)

    arquivosEnviados.sort()
    arquivosRecebidos.sort()

    for i, file in enumerate(arquivosEnviados):
    # LER AQUI O ENVIADO
        readerEnviados = pd.read_csv(pathEnviados + file, sep=";")
    # ESSE AQUI É O NOME DO ARQUIVO RECEBIDO PRA ABIR NA LINHA DE BAIXO
        
        readerRecebidos = pd.read_csv(pathRecebidos + arquivosRecebidos[i], sep=";" , names=['TELEFONE', 'MENSAGEM'])


        enviados_df = readerEnviados.copy()
        recebidos_df = readerRecebidos.copy()


    # ler arquivos
    # enviados_df = pd.read_csv('./enviados.csv', sep=";")
    # Lendo arquivo e nomeando colunas já existentes (criando cabeçalho)
    # recebidos_df = pd.read_csv('./recebidos.csv', sep=";" , names=['TELEFONE', 'MENSAGEM'])

    # renomear colunas do arquivo lido
        enviados_df.rename(columns={
                        'Número': 'TELEFONE', 'status': 'STATUS', 'data de envio': 'DATA'}, inplace=True)

        enviados_df = enviados_df.copy()


        # Remover o 55 dos telefones (enviados)
        for index in enviados_df.index:
            numero = str(enviados_df.loc[index, 'TELEFONE'])

            if numero[0:2] == '55':
                novoNum = numero.lstrip('55')

            enviados_df.at[index, 'TELEFONE'] = novoNum

        # Remover o 55 dos telefones (recebidos)
        for index in recebidos_df.index:
            numero = str(recebidos_df.loc[index, 'TELEFONE'])

            if numero[0:2] == '55':
                novoNum = numero.lstrip('55')

            recebidos_df.at[index, 'TELEFONE'] = novoNum


        # Remover duplicatas
        # Como parâmetro, passar o nome da coluna que deseja remover duplicatas e inplace=True para remover na mesma tabela
        enviados_df.drop_duplicates(subset=['TELEFONE'], inplace=True)

        enviados_df = enviados_df[['TELEFONE', 'DATA', 'STATUS']]


        # remover negativas (erro, inválido, falha - se houver) e alterar e células positivas para “enviado” 
        # for index in enviados_df.index:
        #     if enviados_df.loc[index, 'STATUS'] == 'No':
        #         enviados_df.drop(index, inplace=True, axis=0)

        for index in enviados_df.index:
            newValue = str('Enviado')
            if enviados_df.loc[index, 'STATUS'] == 'Enviada' or enviados_df.loc[index, 'STATUS'] == 'Recebida' or enviados_df.loc[index, 'STATUS'] == 'Lida':
                enviados_df.at[index, 'STATUS'] = newValue



        # Removendo números que estão na blacklist
        numBlacklist = []
        for index in blacklist.index:
            num = str(blacklist.loc[index, 'Telefone'])

            novoNum = num
            blacklist.at[index, 'Telefone'] = novoNum

            numBlacklist.append(blacklist.loc[index, 'Telefone'])
        
        


        for index in enviados_df.index:
            if enviados_df.loc[index, 'TELEFONE'] in numBlacklist:
                enviados_df.drop(index, inplace=True, axis=0)


        # OPENPYXL
        # "Converter" para excel (uso de openpyxl a partir de agora)
        counter += 1
        strCounter = str(counter)
        with pd.ExcelWriter(f'./Conversoes/{strCounter}-{file}.xlsx') as writer:
            enviados_df.to_excel(writer, sheet_name='Enviados', index=False)
            recebidos_df.to_excel(writer, sheet_name='Respostas', index=False)


        # ler arquivo
        teste_wb = load_workbook(f'./Conversoes/{strCounter}-{file}.xlsx')

        # Variável para selecionar a aba
        teste_wb_page = teste_wb['Enviados']
        teste_wb_page2 = teste_wb['Respostas']


        # configurando largura das colunas
        teste_wb_page.column_dimensions['A'].width = 17
        teste_wb_page.column_dimensions['B'].width = 14
        teste_wb_page.column_dimensions['C'].width = 14

        teste_wb_page2.column_dimensions['A'].width = 17
        teste_wb_page2.column_dimensions['B'].width = 300

        # configurando alinhamento do conteúdo das colunas
        for rows in teste_wb_page.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')


        for rows in teste_wb_page2.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')


        # configurando formato de data na coluna DATA

        for col in teste_wb_page.iter_cols(min_row=2, min_col=2, max_col=2):
            for cell in col:
                celula = cell.value  # aqui eu acessei o valor da célula, agora eu posso alterar o valor
                dateSet = celula.split()  # Como isso é uma string (XX/XX/XXXX  HH:MM:SS) e quero só acessar a data, eu faço um split separando a string em uma lista separada por espaço
                justDate = dateSet[0]  # Aqui eu acesso apenas a data
                cell.value = justDate  # Aqui eu altero o valor da célula


        # Bordas das células

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        for col in teste_wb_page.iter_cols(min_row=2, min_col=1, max_col=3):
            for cell in col:
                cell.border = thin_border


        for col in teste_wb_page2.iter_cols(min_row=2, min_col=1, max_col=2):
            for cell in col:
                cell.border = thin_border

        # Deletar colunas (vai variar de arquivo para arquivo)
        teste_wb_page.delete_cols(4, 2)

        # Salvar arquivo
        teste_wb.save(f'{pathSaida}{strCounter}-{file}-result.xlsx')
