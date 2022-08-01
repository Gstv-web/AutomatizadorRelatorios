import pandas as pd
import numpy as np
import xlsxwriter as xlsx
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import os


###  VERIFICAR PATHSAIDA, COUNTER E NA PARTE DO NOME DA CONVERSÃO

path = './Entrada/Bewake/'
pathSaida = './Saida/Bewake/'
blacklist = pd.read_csv('./Blacklist.csv')

def whatsappBewake():
    counter = 0

    for file in os.listdir(path):
        reader = pd.read_csv(path + file, sep=";", encoding="latin-1", on_bad_lines='skip')




        # renomear colunas
        reader.rename(columns={'Celular' : 'TELEFONE', 'Status' : 'STATUS', 'Enviado Em' : 'DATA', 'Resposta' : 'RESPOSTA'}, inplace=True)

        # criar uma cópia para não alterar o original
        enviados_df = reader.copy()
        recebidos_df = reader.copy()


        # Remover (), - e espaços do telefone (enviados)
        for index in enviados_df.index:
            numero = str(enviados_df.loc[index, 'TELEFONE'])

            novoNum = numero.replace("(", "").replace(")", "").replace("-", "").replace(" ", "") # Dá pra usar a mesma função várias vezes na mesma linha
            
            enviados_df.at[index, 'TELEFONE'] = novoNum


        # Remover (), - e espaços do telefone (recebidos)
        for index in recebidos_df.index:
            numero = str(enviados_df.loc[index, 'TELEFONE'])

            novoNum = numero.replace("(", "").replace(")", "").replace("-", "").replace(" ", "") # Dá pra usar a mesma função várias vezes na mesma linha
            
            recebidos_df.at[index, 'TELEFONE'] = novoNum


        # Mostrar somente os telefones da coluna Status que sejam "SUCESSO" ou "RECEBIDO SUCESSO"
        for index in enviados_df.index:
            if enviados_df.loc[index, 'STATUS'] != 'SUCESSO' and enviados_df.loc[index, 'STATUS'] != 'RECEBIDO SUCESSO':
                enviados_df.drop(index, inplace=True, axis=0)

        
        # Alterar células "SUCESSO" e "RECEBIDO SUCESSO" para "ENVIADO"
        for index in enviados_df.index:
            newValue = str('Enviado')
            if enviados_df.loc[index, 'STATUS'] == 'SUCESSO' or enviados_df.loc[index, 'STATUS'] == 'RECEBIDO SUCESSO':
                enviados_df.at[index, 'STATUS'] = newValue

        # Mostrar somente os telefones da coluna Status que sejam "RECEBIDO SUCESSO"
        for index in recebidos_df.index:
            if recebidos_df.loc[index, 'STATUS'] != 'RECEBIDO SUCESSO':
                recebidos_df.drop(index, inplace=True, axis=0)


        # Removendo números que estão na blacklist
        numBlacklist = []
        for index in blacklist.index:
            num = str(blacklist.loc[index, 'Telefone'])

            novoNum = num
            blacklist.at[index, 'Telefone'] = novoNum

            numBlacklist.append(blacklist.loc[index, 'Telefone'])
        

        for index in enviados_df.index:
            if enviados_df.loc[index, 'TELEFONE'] in numBlacklist:
                enviados_df.drop(index, inplace=True, axis=0)
        

        #remover duplicatas
        enviados_df.drop_duplicates(subset=['TELEFONE'], inplace=True)

        enviados_df.drop(columns=['Código', 'Entrada Em', 'Programado Em', 'Nome', 'Tipo', 'Serviço', 'Operadora', 'Whastapp', 'Campanha', 'RESPOSTA', 'Duração'], inplace=True)

        recebidos_df.drop(columns=['Código', 'Entrada Em', 'Programado Em', 'Nome', 'Tipo', 'Serviço', 'STATUS','DATA', 'Operadora', 'Whastapp', 'Campanha', 'Duração'], inplace=True)


        enviados_df = enviados_df[['TELEFONE', 'DATA', 'STATUS']]



        # OPENPYXL
        counter += 1
        strCounter = str(counter)
        with pd.ExcelWriter(f'{strCounter}-{file}.xlsx') as writer:
            enviados_df.to_excel(writer, sheet_name='Enviados', index=False)
            recebidos_df.to_excel(writer, sheet_name='Recebidos', index=False, engine='xlsxwriter')
            

        reader = load_workbook(f'{strCounter}-{file}.xlsx')

        enviados_ws = reader['Enviados']
        recebidos_ws = reader['Recebidos']


        enviados_ws.column_dimensions['A'].width = 17
        enviados_ws.column_dimensions['B'].width = 19
        enviados_ws.column_dimensions['C'].width = 14

        recebidos_ws.column_dimensions['A'].width = 17
        recebidos_ws.column_dimensions['B'].width = 200

        for rows in enviados_ws.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')

        for rows in recebidos_ws.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')


        for col in enviados_ws.iter_cols(min_row=2, min_col=2, max_col=2):
            for cell in col:
                celula = cell.value # aqui eu acessei o valor da célula, agora eu posso alterar o valor
                dateSet = celula.split() # O estilo de data é diferente e precisa ser alterado
                justDate = dateSet[0] # Aqui eu acesso apenas a data
                cell.value = justDate # Aqui eu altero o valor da célula



        # Bordas das células

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        for col in enviados_ws.iter_cols(min_row=2, min_col=1, max_col=3):
            for cell in col:
                cell.border = thin_border


        for col in recebidos_ws.iter_cols(min_row=2, min_col=1, max_col=2):
            for cell in col:
                cell.border = thin_border


        reader.save(f"{pathSaida}{strCounter}-{file}-result.xlsx")
       
        
