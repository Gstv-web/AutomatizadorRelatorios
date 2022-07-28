import pandas as pd
import numpy as np
import xlsxwriter as xlsx
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import os


path = './Entrada/Ligdata/'


def whatsappLigdata():

    for file in os.listdir(path):
        

        readerEnviados = pd.read_excel(path + file, sheet_name='Envio')
        readerRecebidos = pd.read_excel(path + file, sheet_name='Respostas')


        readerEnviados.rename(columns={'Data' : 'DATA', 'Destinatários' : 'TELEFONE', 'Status' : 'STATUS'}, inplace=True)
        readerRecebidos.rename(columns={'Destinatários' : 'TELEFONE', 'Respostas' : 'MENSAGEM'}, inplace=True)


        enviados_df = readerEnviados.copy()
        recebidos_df = readerRecebidos.copy()


        # EDITANDO PLANILHA "ENVIADOS"
        # Removendo 55 do telefone
        for index in enviados_df.index:
            numero = str(enviados_df.loc[index, 'TELEFONE'])

            if numero[0:2] == '55':
                novoNum = numero.lstrip('55')

            enviados_df.at[index, 'TELEFONE'] = novoNum


        enviados_df.drop_duplicates(subset=['TELEFONE'], keep='first', inplace=True)

        enviados_df.drop(columns=['Mensagem'], inplace=True)

        enviados_df = enviados_df[['TELEFONE', 'DATA', 'STATUS']]




        #EDITANDO PLANILHA "RESPOSTAS"
        # Remover 55 do telefone
        for index in recebidos_df.index:
            numero = str(enviados_df.loc[index, 'TELEFONE'])

            if numero[0:2] == '55':
                novoNum = numero.lstrip('55')

            recebidos_df.at[index, 'TELEFONE'] = novoNum


        # remover linhas “no” alterar e células de “yes” para “enviado” 
        for index in enviados_df.index:
            if enviados_df.loc[index, 'STATUS'] == 'Inválida':
                enviados_df.drop(index, inplace=True, axis=0)

        for index in enviados_df.index:
            newValue = str('Enviado')
            if enviados_df.loc[index, 'STATUS'] == 'Enviada':
                enviados_df.at[index, 'STATUS'] = newValue




        # OPENPYXL
        with pd.ExcelWriter('./Conversoes/RelatorioWhatsapp-Ligdata.xlsx') as writer:
            enviados_df.to_excel(writer, sheet_name='Enviados', index=False)
            recebidos_df.to_excel(writer, sheet_name='Respostas', index=False)


        reader = load_workbook('./Conversoes/RelatorioWhatsapp-Ligdata.xlsx')


        enviados_ws = reader['Enviados']
        recebidos_ws = reader['Respostas']

        enviados_ws.column_dimensions['A'].width = 17
        enviados_ws.column_dimensions['B'].width = 14
        enviados_ws.column_dimensions['C'].width = 14

        recebidos_ws.column_dimensions['A'].width = 17
        recebidos_ws.column_dimensions['B'].width = 200

        for rows in enviados_ws.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')

        for rows in recebidos_ws.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')


        for col in enviados_ws.iter_cols(min_row=2, min_col=2, max_col=2):
            for cell in col:
                celula = cell.value.strftime("%d/%m/%Y") # aqui eu acessei o valor da célula, agora eu posso alterar o valor
                dateSet = celula.split() # O estilo de data é diferente e precisa ser alterado
                justDate = dateSet[0] # Aqui eu acesso apenas a data
                cell.value = justDate # Aqui eu altero o valor da célula


        # Bordas das células

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        for col in enviados_ws.iter_cols(min_row=2, min_col=1, max_col=3):
            for cell in col:
                cell.border = thin_border


        for col in recebidos_ws.iter_cols(min_row=2, min_col=1, max_col=2):
            for cell in col:
                cell.border = thin_border


        reader.save("./Saida/Ligdata/RelatorioWhatsapp-Ligdata.xlsx")